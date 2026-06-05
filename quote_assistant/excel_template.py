from __future__ import annotations

import hashlib
import re
from dataclasses import dataclass
from pathlib import Path
from typing import Any

try:
    from copy import copy

    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image as OpenpyxlImage
except ImportError:  # pragma: no cover - validated at runtime.
    copy = None  # type: ignore[assignment]
    load_workbook = None  # type: ignore[assignment]
    OpenpyxlImage = None  # type: ignore[assignment]

try:
    from PIL import Image as PILImage
    from PIL import ImageOps
except ImportError:  # pragma: no cover - validated at runtime.
    PILImage = None  # type: ignore[assignment]
    ImageOps = None  # type: ignore[assignment]


PROJECT_ROOT = Path(__file__).resolve().parents[1]
TEMPLATES_DIR = PROJECT_ROOT / "templates"
SHEET_METAL_TEMPLATE = TEMPLATES_DIR / "sheet_metal_cost_template.xlsx"
SHEET_NAME = "成本分析"
DATA_START_ROW = 3
TEMPLATE_LAST_DATA_ROW = 22
LAST_COLUMN = "CO"
IMAGE_COLUMN = "F"
IMAGE_COLUMN_WIDTH = 24
IMAGE_ROW_HEIGHT = 82
IMAGE_DISPLAY_MAX_WIDTH_PX = 160
IMAGE_DISPLAY_MAX_HEIGHT_PX = 100
IMAGE_EMBED_MAX_SIZE = (900, 600)


@dataclass(frozen=True)
class ExcelBuildResult:
    path: Path
    row_count: int
    warnings: list[str]
    image_count: int = 0


INPUT_COLUMNS = {
    "no": "A",
    "level": "B",
    "part_number": "C",
    "has_children": "D",
    "product_type": "E",
    "drawing_ref": "F",
    "remark": "G",
    "description": "H",
    "qty": "I",
    "bom_price": "K",
    "material_drawing": "L",
    "material_substitute": "M",
    "raw_weight_kg": "N",
    "net_weight_kg": "O",
    "material_unit_price": "P",
    "laser_cut_unit_price": "S",
    "laser_cut_length_m": "T",
    "laser_hole_unit_price": "V",
    "laser_hole_length_m": "W",
    "blanking_other_process_name": "Y",
    "blanking_other_unit_price": "Z",
    "blanking_other_qty": "AA",
    "chamfer_unit_price": "AE",
    "chamfer_qty": "AF",
    "tapping_unit_price": "AH",
    "tapping_qty": "AI",
    "polishing_unit_price": "AK",
    "polishing_area_m2": "AL",
    "bend_unit_price": "AN",
    "bend_count": "AO",
    "edge_trim_unit_price": "AQ",
    "edge_trim_hours": "AR",
    "milling_unit_price": "AT",
    "milling_hours": "AU",
    "brushing_unit_price": "AW",
    "brushing_area_m2": "AX",
    "punching_unit_price": "AZ",
    "punching_qty": "BA",
    "rivet_unit_price": "BC",
    "rivet_qty": "BD",
    "welding_unit_price": "BF",
    "welding_hours": "BG",
    "other_process_name": "BI",
    "other_process_unit_price": "BJ",
    "other_process_qty": "BK",
    "plating_unit_price": "BO",
    "plating_weight_kg": "BP",
    "spraying_unit_price": "BR",
    "spraying_area_dm2": "BS",
    "hot_dip_zinc_unit_price": "BU",
    "hot_dip_zinc_qty": "BV",
    "zinc_repair_unit_price": "BX",
    "zinc_repair_hours": "BY",
    "surface_process_name": "CA",
    "surface_unit_price": "CB",
    "surface_qty": "CC",
    "packing_cost": "CI",
    "shipping_cost": "CJ",
    "note": "CO",
}


FORMULA_COLUMNS = {
    "Q": '=IF(OR(N{r}="",P{r}=""),"",N{r}*P{r})',
    "R": '=IFERROR(Q{r}/CN{r},"")',
    "U": '=IF(OR(T{r}="",S{r}=""),"",T{r}*S{r})',
    "X": '=IF(OR(V{r}="",W{r}=""),"",V{r}*W{r})',
    "AB": '=IF(OR(Z{r}="",AA{r}=""),"",Z{r}*AA{r})',
    "AC": '=IF(COUNT(U{r}:AB{r})=0,"",SUM(U{r}:AB{r}))',
    "AG": '=IF(OR(AE{r}="",AF{r}=""),"",AE{r}*AF{r})',
    "AJ": '=IF(OR(AH{r}="",AI{r}=""),"",AH{r}*AI{r})',
    "AM": '=IF(OR(AK{r}="",AL{r}=""),"",AK{r}*AL{r})',
    "AP": '=IF(OR(AN{r}="",AO{r}=""),"",AN{r}*AO{r})',
    "AS": '=IF(OR(AQ{r}="",AR{r}=""),"",AQ{r}*AR{r})',
    "AV": '=IF(OR(AT{r}="",AU{r}=""),"",AT{r}*AU{r})',
    "AY": '=IF(OR(AW{r}="",AX{r}=""),"",AW{r}*AX{r})',
    "BB": '=IF(OR(AZ{r}="",BA{r}=""),"",AZ{r}*BA{r})',
    "BE": '=IF(OR(BC{r}="",BD{r}=""),"",BC{r}*BD{r})',
    "BH": '=IF(OR(BF{r}="",BG{r}=""),"",BF{r}*BG{r})',
    "BL": '=IF(OR(BJ{r}="",BK{r}=""),"",BJ{r}*BK{r})',
    "BM": '=IF(COUNT(AG{r},AJ{r},AM{r},AP{r},AS{r},AV{r},AY{r},BB{r},BE{r},BH{r},BL{r})=0,"",SUM(AG{r},AJ{r},AM{r},AP{r},AS{r},AV{r},AY{r},BB{r},BE{r},BH{r},BL{r}))',
    "BQ": '=IF(OR(BO{r}="",BP{r}=""),"",BO{r}*BP{r})',
    "BT": '=IF(OR(BR{r}="",BS{r}=""),"",BR{r}*BS{r})',
    "BW": '=IF(OR(BU{r}="",BV{r}=""),"",BU{r}*BV{r})',
    "BZ": '=IF(OR(BX{r}="",BY{r}=""),"",BX{r}*BY{r})',
    "CD": '=IF(OR(CB{r}="",CC{r}=""),"",CB{r}*CC{r})',
    "CE": '=IF(COUNT(BQ{r},BT{r},BW{r},BZ{r},CD{r})=0,"",SUM(BQ{r},BT{r},BW{r},BZ{r},CD{r}))',
    "CG": '=IF(COUNT(AC{r},BM{r},CE{r})=0,"",SUM(AC{r},BM{r},CE{r}))',
    "CH": '=IF(COUNT(CG{r},Q{r})=0,"",SUM(CG{r},Q{r}))',
    "CK": '=IF(COUNT(CG{r}:CJ{r})=0,"",SUM(CG{r}:CJ{r})*0.05)',
    "CL": '=IF(COUNT(CG{r}:CK{r})=0,"",SUM(CG{r}:CK{r})*0.1)',
    "CN": '=IF(COUNT(CH{r}:CL{r})=0,"",SUM(CL{r},CK{r},CH{r},CJ{r},CI{r}))',
}


DIRECT_COST_KEYS = {
    "raw_weight_kg",
    "material_unit_price",
    "laser_cut_length_m",
    "laser_hole_length_m",
    "blanking_other_qty",
    "chamfer_qty",
    "tapping_qty",
    "polishing_area_m2",
    "bend_count",
    "edge_trim_hours",
    "milling_hours",
    "brushing_area_m2",
    "punching_qty",
    "rivet_qty",
    "welding_hours",
    "other_process_qty",
    "plating_weight_kg",
    "spraying_area_dm2",
    "hot_dip_zinc_qty",
    "zinc_repair_hours",
    "surface_qty",
    "packing_cost",
    "shipping_cost",
}

PRICE_INPUT_KEYS = {
    "unit_price",
    "material_unit_price",
    "laser_cut_unit_price",
    "laser_hole_unit_price",
    "blanking_other_unit_price",
    "chamfer_unit_price",
    "tapping_unit_price",
    "polishing_unit_price",
    "bend_unit_price",
    "edge_trim_unit_price",
    "milling_unit_price",
    "brushing_unit_price",
    "punching_unit_price",
    "rivet_unit_price",
    "welding_unit_price",
    "other_process_unit_price",
    "plating_unit_price",
    "spraying_unit_price",
    "hot_dip_zinc_unit_price",
    "zinc_repair_unit_price",
    "surface_unit_price",
    "packing_cost",
    "shipping_cost",
}


def _is_blank(value: Any) -> bool:
    return value is None or value == "" or value == []


def _normalize_yes_no(value: Any) -> str:
    text = str(value or "").strip().lower()
    if text in {"y", "yes", "true", "1", "有", "是"}:
        return "Y"
    if text in {"n", "no", "false", "0", "无", "否"}:
        return "N"
    return str(value or "").strip()


def _normalize_number(value: Any) -> Any:
    if _is_blank(value):
        return None
    if isinstance(value, (int, float)):
        return value
    text = str(value).strip()
    if text in {"未识别", "待确认", "图纸未标注", "unknown", "Unknown", "N/A", "-"}:
        return None
    try:
        return float(text.replace(",", ""))
    except ValueError:
        return value


def _truthy_cost(value: Any) -> bool:
    normalized = _normalize_number(value)
    if normalized is None:
        return False
    if isinstance(normalized, (int, float)) and normalized == 0:
        return False
    return True


def _has_direct_cost(row: dict[str, Any]) -> bool:
    return any(_truthy_cost(row.get(key)) for key in DIRECT_COST_KEYS)


def _has_priced_cost(row: dict[str, Any]) -> bool:
    return any(_truthy_cost(row.get(key)) for key in PRICE_INPUT_KEYS)


def _has_priced_subtree(rows: list[dict[str, Any]], index: int) -> bool:
    current_level = int(_normalize_number(rows[index].get("level")) or 0)
    if _has_priced_cost(rows[index]):
        return True
    for child_index in range(index + 1, len(rows)):
        child_level = int(_normalize_number(rows[child_index].get("level")) or 0)
        if child_level <= current_level:
            break
        if _has_priced_cost(rows[child_index]):
            return True
    return False


def _truthy_flag(value: Any) -> bool:
    return str(value or "").strip().lower() in {"1", "true", "yes", "y", "on"}


def _normalize_match_text(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").lower())


def _row_image_tokens(row: dict[str, Any]) -> list[str]:
    tokens: set[str] = set()
    for key in ("drawing_ref", "part_number", "description", "note"):
        text = str(row.get(key) or "")
        for chunk in re.findall(r"[A-Za-z0-9][A-Za-z0-9._-]{2,}", text):
            normalized = _normalize_match_text(chunk)
            if len(normalized) >= 4:
                tokens.add(normalized)
            for part in re.split(r"[-_.]+", chunk):
                normalized_part = _normalize_match_text(part)
                if len(normalized_part) >= 4:
                    tokens.add(normalized_part)
    return sorted(tokens, key=len, reverse=True)


def _image_candidates(image_assets: list[dict[str, Any]] | None) -> list[dict[str, Any]]:
    candidates: list[dict[str, Any]] = []
    for asset in image_assets or []:
        if not isinstance(asset, dict) or asset.get("type") != "image":
            continue
        raw_path = str(asset.get("path") or "")
        if not raw_path:
            continue
        path = Path(raw_path)
        if not path.exists():
            continue
        label = str(asset.get("label") or path.name)
        search_key = _normalize_match_text(
            " ".join(
                [
                    label,
                    str(asset.get("kind") or ""),
                    str(asset.get("region") or ""),
                    str(asset.get("source") or ""),
                    path.name,
                    path.stem,
                    path.parent.name,
                ]
            )
        )
        candidates.append(
            {
                "path": path,
                "label": label,
                "search_key": search_key,
                "kind": asset.get("kind"),
                "row_no": asset.get("row_no"),
            }
        )
    return candidates


def _select_row_image(
    row: dict[str, Any],
    candidates: list[dict[str, Any]],
) -> dict[str, Any] | None:
    image_policy = str(row.get("image_policy") or "").strip().lower()
    if _truthy_flag(row.get("suppress_image")) or image_policy in {"none", "blank", "suppress"}:
        return None
    if not candidates:
        return None

    row_no = _normalize_number(row.get("no"))
    if isinstance(row_no, (int, float)):
        for candidate in candidates:
            candidate_row_no = _normalize_number(candidate.get("row_no"))
            if isinstance(candidate_row_no, (int, float)) and int(candidate_row_no) == int(row_no):
                return candidate

    tokens = _row_image_tokens(row)
    best_candidate: dict[str, Any] | None = None
    best_score: tuple[int, int] = (0, 0)
    for candidate in candidates:
        search_key = str(candidate.get("search_key") or "")
        match_score = max((len(token) for token in tokens if token in search_key), default=0)
        priority = 2 if candidate.get("kind") == "drawing_region" else 1
        score = (match_score, priority) if match_score > 0 else (0, 0)
        if score > best_score:
            best_score = score
            best_candidate = candidate

    if best_candidate:
        return best_candidate
    if len(candidates) == 1:
        return candidates[0]
    return None


def _thumbnail_path(output_path: Path, source_path: Path) -> Path:
    stem = re.sub(r"[^A-Za-z0-9_.-]+", "_", source_path.stem).strip("._") or "image"
    digest = hashlib.sha1(str(source_path).encode("utf-8")).hexdigest()[:10]
    return output_path.parent / "_excel_images" / f"{stem[:48]}-{digest}.png"


def _prepare_thumbnail(source_path: Path, output_path: Path) -> Path:
    if PILImage is None or ImageOps is None:
        raise RuntimeError("Excel image output requires Pillow. Install project requirements first.")

    target_path = _thumbnail_path(output_path, source_path)
    if target_path.exists() and target_path.stat().st_mtime >= source_path.stat().st_mtime:
        return target_path

    target_path.parent.mkdir(parents=True, exist_ok=True)
    with PILImage.open(source_path) as image:
        image = ImageOps.exif_transpose(image)
        image.thumbnail(IMAGE_EMBED_MAX_SIZE)
        if image.mode in {"RGBA", "LA"}:
            background = PILImage.new("RGB", image.size, "white")
            alpha = image.getchannel("A")
            background.paste(image.convert("RGBA"), mask=alpha)
            image = background
        elif image.mode != "RGB":
            image = image.convert("RGB")
        image.save(target_path, "PNG", optimize=True)
    return target_path


def _fit_dimensions(width: int, height: int) -> tuple[int, int]:
    if width <= 0 or height <= 0:
        return IMAGE_DISPLAY_MAX_WIDTH_PX, IMAGE_DISPLAY_MAX_HEIGHT_PX
    scale = min(IMAGE_DISPLAY_MAX_WIDTH_PX / width, IMAGE_DISPLAY_MAX_HEIGHT_PX / height, 1)
    return max(1, int(width * scale)), max(1, int(height * scale))


def _insert_row_image(ws: Any, excel_row: int, image_path: Path, output_path: Path) -> None:
    if OpenpyxlImage is None:
        raise RuntimeError("Excel image output requires openpyxl image support.")

    thumbnail = _prepare_thumbnail(image_path, output_path)
    image = OpenpyxlImage(str(thumbnail))
    image.width, image.height = _fit_dimensions(int(image.width), int(image.height))
    ws[f"{IMAGE_COLUMN}{excel_row}"] = None
    ws.add_image(image, f"{IMAGE_COLUMN}{excel_row}")
    current_width = ws.column_dimensions[IMAGE_COLUMN].width or 0
    ws.column_dimensions[IMAGE_COLUMN].width = max(current_width, IMAGE_COLUMN_WIDTH)
    current_height = ws.row_dimensions[excel_row].height or 0
    ws.row_dimensions[excel_row].height = max(current_height, IMAGE_ROW_HEIGHT)


def _copy_row_format(ws: Any, source_row: int, target_row: int) -> None:
    if copy is None:
        return
    for col in range(1, ws.max_column + 1):
        source = ws.cell(source_row, col)
        target = ws.cell(target_row, col)
        if source.has_style:
            target._style = copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy(source.alignment)
        if source.border:
            target.border = copy(source.border)
        if source.fill:
            target.fill = copy(source.fill)
        if source.font:
            target.font = copy(source.font)


def _image_anchor_row(image: Any) -> int | None:
    marker = getattr(getattr(image, "anchor", None), "_from", None)
    if marker is None:
        return None
    return int(marker.row) + 1


def _clear_data_row_images(ws: Any, first_row: int, last_row: int) -> None:
    if not hasattr(ws, "_images"):
        return
    ws._images = [
        image
        for image in ws._images
        if not (
            (anchor_row := _image_anchor_row(image)) is not None
            and first_row <= anchor_row <= last_row
        )
    ]


def _prepare_rows(ws: Any, row_count: int) -> None:
    needed_last_row = DATA_START_ROW + max(row_count, 1) - 1
    if needed_last_row > TEMPLATE_LAST_DATA_ROW:
        insert_count = needed_last_row - TEMPLATE_LAST_DATA_ROW
        ws.insert_rows(TEMPLATE_LAST_DATA_ROW + 1, insert_count)
        for row in range(TEMPLATE_LAST_DATA_ROW + 1, needed_last_row + 1):
            _copy_row_format(ws, TEMPLATE_LAST_DATA_ROW, row)

    clear_last_row = max(TEMPLATE_LAST_DATA_ROW, needed_last_row)
    _clear_data_row_images(ws, DATA_START_ROW, clear_last_row)

    for row in range(DATA_START_ROW, clear_last_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row, col)
            cell.value = None
            cell.hyperlink = None
            cell.comment = None


def _direct_child_rows(rows: list[dict[str, Any]], index: int) -> list[int]:
    current_level = int(_normalize_number(rows[index].get("level")) or 0)
    children: list[int] = []
    for child_index in range(index + 1, len(rows)):
        child_level = int(_normalize_number(rows[child_index].get("level")) or 0)
        if child_level <= current_level:
            break
        if child_level == current_level + 1:
            children.append(DATA_START_ROW + child_index)
    return children


def _write_j_formula(ws: Any, excel_row: int, rows: list[dict[str, Any]], index: int) -> None:
    row = rows[index]
    has_children = _normalize_yes_no(row.get("has_children"))
    child_rows = _direct_child_rows(rows, index)
    explicit_unit_price = _normalize_number(row.get("unit_price"))

    if explicit_unit_price is not None:
        ws[f"J{excel_row}"] = explicit_unit_price
        return

    if has_children == "Y" and child_rows:
        child_refs = ",".join(f"K{child_row}" for child_row in child_rows)
        if not _has_priced_subtree(rows, index):
            ws[f"J{excel_row}"] = None
            return
        ws[f"J{excel_row}"] = f'=IF(COUNT({child_refs},CN{excel_row})=0,"",SUM({child_refs},CN{excel_row}))'
        return

    if _has_priced_cost(row):
        ws[f"J{excel_row}"] = f'=IF(CN{excel_row}="","",CN{excel_row})'
        return

    ws[f"J{excel_row}"] = None


def _write_row(ws: Any, excel_row: int, rows: list[dict[str, Any]], index: int) -> None:
    row = rows[index]
    normalized = dict(row)
    normalized.setdefault("no", index + 1)
    normalized.setdefault("qty", 1)
    normalized["has_children"] = _normalize_yes_no(normalized.get("has_children"))

    for key, column in INPUT_COLUMNS.items():
        value = normalized.get(key)
        if key in {
            "qty",
            "bom_price",
            "raw_weight_kg",
            "net_weight_kg",
            "material_unit_price",
            "laser_cut_unit_price",
            "laser_cut_length_m",
            "laser_hole_unit_price",
            "laser_hole_length_m",
            "blanking_other_unit_price",
            "blanking_other_qty",
            "chamfer_unit_price",
            "chamfer_qty",
            "tapping_unit_price",
            "tapping_qty",
            "polishing_unit_price",
            "polishing_area_m2",
            "bend_unit_price",
            "bend_count",
            "edge_trim_unit_price",
            "edge_trim_hours",
            "milling_unit_price",
            "milling_hours",
            "brushing_unit_price",
            "brushing_area_m2",
            "punching_unit_price",
            "punching_qty",
            "rivet_unit_price",
            "rivet_qty",
            "welding_unit_price",
            "welding_hours",
            "other_process_unit_price",
            "other_process_qty",
            "plating_unit_price",
            "plating_weight_kg",
            "spraying_unit_price",
            "spraying_area_dm2",
            "hot_dip_zinc_unit_price",
            "hot_dip_zinc_qty",
            "zinc_repair_unit_price",
            "zinc_repair_hours",
            "surface_unit_price",
            "surface_qty",
            "packing_cost",
            "shipping_cost",
        }:
            value = _normalize_number(value)
        if not _is_blank(value):
            ws[f"{column}{excel_row}"] = value

    _write_j_formula(ws, excel_row, rows, index)
    ws[f"K{excel_row}"] = f'=IF(J{excel_row}="","",J{excel_row}*I{excel_row})'

    for column, formula in FORMULA_COLUMNS.items():
        ws[f"{column}{excel_row}"] = formula.format(r=excel_row)


def build_sheet_metal_workbook(
    payload: dict[str, Any],
    output_path: Path,
    image_assets: list[dict[str, Any]] | None = None,
) -> ExcelBuildResult:
    if load_workbook is None:
        raise RuntimeError("Excel output requires openpyxl. Install project requirements first.")
    if not SHEET_METAL_TEMPLATE.exists():
        raise FileNotFoundError(f"Missing Excel template: {SHEET_METAL_TEMPLATE}")

    rows = payload.get("rows") if isinstance(payload, dict) else None
    if not isinstance(rows, list) or not rows:
        raise ValueError("Excel payload must contain a non-empty rows array.")
    row_payloads = [row for row in rows if isinstance(row, dict)]
    if not row_payloads:
        raise ValueError("Excel payload rows must be objects.")

    workbook = load_workbook(SHEET_METAL_TEMPLATE)
    ws = workbook[SHEET_NAME] if SHEET_NAME in workbook.sheetnames else workbook.active
    _prepare_rows(ws, len(row_payloads))

    warnings: list[str] = []
    candidates = _image_candidates(image_assets)
    inserted_images = 0
    for index, row in enumerate(row_payloads):
        excel_row = DATA_START_ROW + index
        if not row.get("part_number"):
            warnings.append(f"Row {index + 1} is missing part_number.")
        _write_row(ws, excel_row, row_payloads, index)
        image_candidate = _select_row_image(row, candidates)
        if image_candidate:
            try:
                _insert_row_image(ws, excel_row, image_candidate["path"], output_path)
                inserted_images += 1
            except Exception as exc:
                warnings.append(f"Row {index + 1} image insert failed: {exc}")

    if hasattr(workbook, "calculation"):
        workbook.calculation.fullCalcOnLoad = True
        workbook.calculation.forceFullCalc = True

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
    return ExcelBuildResult(
        path=output_path,
        row_count=len(row_payloads),
        warnings=warnings,
        image_count=inserted_images,
    )
